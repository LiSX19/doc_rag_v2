"""
Excel文档加载器

支持.xlsx和.xls格式。
"""

from pathlib import Path
from typing import Any, Dict, List, Optional, Union

from .base import BaseLoader


class ExcelLoader(BaseLoader):
    """Excel加载器"""
    
    def __init__(self, config: Optional[Dict[str, Any]] = None):
        """
        初始化Excel加载器
        
        Args:
            config: 配置字典
        """
        super().__init__(config)
        
        self._unstructured_available = self._check_unstructured()
        self._openpyxl_available = self._check_openpyxl()
        self._xlrd_available = self._check_xlrd()
    
    def _check_unstructured(self) -> bool:
        """检查Unstructured是否可用"""
        try:
            from unstructured.partition.xlsx import partition_xlsx
            from unstructured.partition.xls import partition_xls
            return True
        except ImportError:
            return False
    
    def _check_openpyxl(self) -> bool:
        """检查openpyxl是否可用"""
        try:
            import openpyxl
            return True
        except ImportError:
            return False
    
    def _check_xlrd(self) -> bool:
        """检查xlrd是否可用"""
        try:
            import xlrd
            return True
        except ImportError:
            return False
    
    def supports(self, file_path: Union[str, Path]) -> bool:
        """检查是否支持该文件"""
        ext = Path(file_path).suffix.lower()
        return ext in ['.xlsx', '.xls']
    
    def load(self, file_path: Union[str, Path]) -> Dict[str, Any]:
        """
        加载Excel文档
        
        Args:
            file_path: Excel文件路径
            
        Returns:
            包含文档内容的字典
        """
        file_path = Path(file_path)
        
        if not file_path.exists():
            raise FileNotFoundError(f"文件不存在: {file_path}")
        
        ext = file_path.suffix.lower()
        
        # 尝试使用Unstructured
        if self._unstructured_available:
            try:
                if ext == '.xlsx':
                    return self._load_with_unstructured_xlsx(file_path)
                else:
                    return self._load_with_unstructured_xls(file_path)
            except Exception as e:
                pass
        
        # 使用专用库
        if ext == '.xlsx' and self._openpyxl_available:
            return self._load_with_openpyxl(file_path)
        elif ext == '.xls' and self._xlrd_available:
            return self._load_with_xlrd(file_path)
        
        raise RuntimeError(f"无法解析Excel文件: {file_path}")
    
    def _load_with_unstructured_xlsx(self, file_path: Path) -> Dict[str, Any]:
        """使用Unstructured加载.xlsx"""
        from unstructured.partition.xlsx import partition_xlsx
        
        # 获取语言配置
        languages = self.config.get('loader.unstructured.languages', ['chi_sim', 'eng'])
        
        elements = partition_xlsx(filename=str(file_path), languages=languages)
        
        texts = []
        for element in elements:
            text = str(element)
            if text.strip():
                texts.append(text)
        
        content = '\n\n'.join(texts)
        
        metadata = self.extract_metadata(file_path)
        metadata['parser'] = 'unstructured_xlsx'
        
        return {
            'content': content,
            'metadata': metadata,
            'pages': [],
        }
    
    def _load_with_unstructured_xls(self, file_path: Path) -> Dict[str, Any]:
        """使用Unstructured加载.xls"""
        from unstructured.partition.xls import partition_xls
        
        # 获取语言配置
        languages = self.config.get('loader.unstructured.languages', ['chi_sim', 'eng'])
        
        elements = partition_xls(filename=str(file_path), languages=languages)
        
        texts = []
        for element in elements:
            text = str(element)
            if text.strip():
                texts.append(text)
        
        content = '\n\n'.join(texts)
        
        metadata = self.extract_metadata(file_path)
        metadata['parser'] = 'unstructured_xls'
        
        return {
            'content': content,
            'metadata': metadata,
            'pages': [],
        }
    
    def _load_with_openpyxl(self, file_path: Path) -> Dict[str, Any]:
        """使用openpyxl加载.xlsx"""
        import openpyxl
        
        wb = openpyxl.load_workbook(str(file_path), data_only=True)
        
        texts = []
        sheet_names = []
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            sheet_names.append(sheet_name)
            
            texts.append(f"=== Sheet: {sheet_name} ===")
            
            for row in sheet.iter_rows():
                row_texts = []
                for cell in row:
                    if cell.value is not None:
                        row_texts.append(str(cell.value))
                if row_texts:
                    texts.append(' | '.join(row_texts))
        
        content = '\n'.join(texts)
        
        metadata = self.extract_metadata(file_path)
        metadata['parser'] = 'openpyxl'
        metadata['sheet_count'] = len(sheet_names)
        metadata['sheet_names'] = sheet_names
        
        return {
            'content': content,
            'metadata': metadata,
            'pages': [],
        }
    
    def _load_with_xlrd(self, file_path: Path) -> Dict[str, Any]:
        """使用xlrd加载.xls"""
        import xlrd
        
        wb = xlrd.open_workbook(str(file_path))
        
        texts = []
        sheet_names = wb.sheet_names()
        
        for sheet_name in sheet_names:
            sheet = wb.sheet_by_name(sheet_name)
            
            texts.append(f"=== Sheet: {sheet_name} ===")
            
            for row_idx in range(sheet.nrows):
                row_values = []
                for col_idx in range(sheet.ncols):
                    cell_value = sheet.cell_value(row_idx, col_idx)
                    if cell_value:
                        row_values.append(str(cell_value))
                if row_values:
                    texts.append(' | '.join(row_values))
        
        content = '\n'.join(texts)
        
        metadata = self.extract_metadata(file_path)
        metadata['parser'] = 'xlrd'
        metadata['sheet_count'] = len(sheet_names)
        metadata['sheet_names'] = sheet_names
        
        return {
            'content': content,
            'metadata': metadata,
            'pages': [],
        }
